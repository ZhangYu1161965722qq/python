import re
import openpyxl
import os
import traceback
from tkinter import Tk,Frame,Text,Scrollbar,EventType,filedialog
from tkinter.ttk import Button
import time


def createRegExp():
    wb=openpyxl.load_workbook('config.xlsx')
    ws_mark=wb['手动配置']
    ws_regExp=wb['Regular Expression-自动生成']

    ws_regExp.delete_rows(2, ws_regExp.max_row-1)

    max_row=ws_mark.max_row

    list_regExp=[]
    for i in range(2,max_row+1):
        str_mark=ws_mark.cell(i,1).value

        if  str_mark is None:continue
        if str_mark.strip()=='':continue

        list_re=[]
        str_re=regExpReplace(str_mark,1)
        ws_regExp.cell(i,1).value=str_re
        list_re.append(str_re)

        str_re=regExpReplace(str_mark,2)
        ws_regExp.cell(i,2).value=str_re
        list_re.append(str_re)

        list_regExp.append(list_re)

    wb.save('config.xlsx')

    return list_regExp


def regExpReplace(str_mark,tag):
    if tag ==1:
        str_replace1=''
        str_replace2=''
    elif tag==2:
        str_replace1='('
        str_replace2=')'

    str_re=re.sub('【\s*@\s*',str_replace1,str_mark)
    str_re=re.sub('\s*@\s*】',str_replace2,str_re)

    str_re=re.sub('【','',str_re)
    str_re=re.sub('】','',str_re)

    str_re=re.sub('\s*','',str_re)
    return str_re


# # 递归方法
# def listFilePath(dict_Path,str_path):
#     dict_Path[str_path]=[]
#     # 取出所有文件路径
#     list_path= os.listdir(str_path)
    
#     for p in list_path:
#         print(p)
#         path_sub=os.path.join(str_path,p)
#         if os.path.isdir(path_sub):
#             listFilePath(dict_Path,path_sub)
#         else :
#             # isfile判断不了，有问题
#             dict_Path[str_path].append(path_sub)


# # 遍历文件夹及其子目录，来源于网络
# def getAllPath(dir_path):
#     dict_Path={}
#     for root, dirs, files in os.walk(dir_path):
#         # root当前遍历到的文件夹路径
#         dict_Path[root]=[]

#         # dirs当前文件夹中的所有文件夹名
#         for dir in dirs:
#             dict_Path[os.path.join(root, dir)]=[]

#         # files当前文件夹中的所有文件名
#         for file in files:
#             dict_Path[root].append(os.path.join(root, file))
#     return dict_Path


def setCell(myWorksheet,row_cell,num_columns,list_value):
    for i in range(num_columns):
        myCell = myWorksheet.cell(row_cell,i+1)
        myCell.value = list_value[i]      # 赋值
        myCell.alignment = openpyxl.styles.Alignment(wrap_text=True)    # 自动换行


def setTableTitle(myWorksheet,num_columns):
    fille = openpyxl.styles.PatternFill('solid', fgColor="C0C0C0")  # 填充色
    side = openpyxl.styles.Side(style='thin', color='000000')   # 表格线样式

    for i in range(num_columns):
        myWorksheet.cell(1,i+1).border=side
        myWorksheet.cell(1,i+1).fill=fille
        c_letter=openpyxl.utils.get_column_letter(i+1)  # 列数字号变字母
        myWorksheet.column_dimensions[c_letter].width=40 # 设置列宽


def getStr_source(str_path):
    str_source=''
    for charset in ['utf-8','GB2312','GBK','GB18030','hz']:
        try:
            with open(str_path,'r',encoding=charset) as f:
                str_source=f.read()
            break
        except UnicodeDecodeError:
            # print(charset)
            continue

    return str_source


def runRegExpFindAll(dir_path,outputFilename):
    list_regExp=createRegExp()

    dict_result_all={}

    wb=openpyxl.Workbook()
    ws_reuslt=wb.create_sheet('明细')

    ws_reuslt.column_dimensions['a'].width=40 # 设置列宽

    setCell(ws_reuslt,1,4,['路径','文件夹','操作类型','操作对象名'])

    setTableTitle(ws_reuslt,4)

    i=2

    # root当前遍历到的文件夹路径；dirs当前文件夹中的所有文件夹名；files当前文件夹中的所有文件名
    for root, dirs, files in os.walk(dir_path):
        # 在所有文件路径循环取出文本内容，用正则表达式匹配
        for file in files:
            str_path=os.path.join(root,file)

            if str_path.endswith('python_zhangyu.txt'):continue

            foldername=root[root.rfind('\\')+1:]

            str_source=getStr_source(str_path)

            if str_source.strip()=='':continue

            # 在正则表达式匹配模式中循环
            for list_re in list_regExp:
                # 正则表达式查找所有匹配的字符串
                list_resut_regExp=re.findall(pattern=list_re[0],string=str_source,flags=re.IGNORECASE)

                # 匹配list不为空
                if list_resut_regExp:
                    for n in range(len(list_resut_regExp)):
                        list_resut_regExp2=re.search(pattern=list_re[1],string=list_resut_regExp[n],flags=re.IGNORECASE)

                        # 赋值
                        setCell(ws_reuslt,i,4,[str_path,foldername,list_resut_regExp[n],list_resut_regExp2.group(1)])

                        i+=1

                        # 文件夹.操作.对象
                        key_all='%s\\%s\\%s' %(foldername , list_resut_regExp[n].split(' ')[0],list_resut_regExp2.group(1))
                        key_all=key_all.lower()
                        if key_all not in dict_result_all:
                            dict_result_all[key_all]=[list_resut_regExp[n],list_resut_regExp2.group(1)]

    # 汇总去重
    ws_all=wb._sheets[0]    # 汇总表
    ws_all.title='汇总'

    setCell(ws_all,1,3,['文件夹','操作类型','操作对象名'])

    setTableTitle(ws_all,3)

    i=2
    for k,v in dict_result_all.items():
        # 赋值
        setCell(ws_all,i,3,[k[:k.find('\\')],v[0],v[1]])
        i+=1

    wb.save(outputFilename)

# 函数：窗口初始化
def windowInit():
    mainWindow=Tk()
    mainWindow.title('脚本分类、合并')

    mainWindow.attributes('-topmost',True)  # 窗口置顶

    bgColor= 'white'
    mainWindow.config(background=bgColor,padx=10,pady=10)

    width_win =700
    height_win = 350
    x_win = (mainWindow.winfo_screenwidth() // 2) - (width_win // 2)
    y_win = (mainWindow.winfo_screenheight() // 3) - (height_win // 3)

    # 窗口居中，设置 窗口大小、位置：字符串格式：width x height + x + y
    mainWindow.geometry('{}x{}+{}+{}'.format(width_win, height_win, x_win, y_win))

    myFrame=Frame(mainWindow,bg=bgColor)

    tip_txt_path='点击“选择文件夹”按钮，或输入文件夹路径'

    outputFilename='py-生成-脚本分类.xlsx'
    # 运行信息框
    str_info='''     ---------------------------------------说明：---------------------------------------
     功能：按标记查找文件夹下所有文本内容记录到excel；合并同一文件夹下的文本内容到一个文件

     一、运行前准备（若配置不变，可不用修改；注意：不要更改config.xlsx文件名、表名及结构）
        配置同文件夹下excel文件（config.xlsx）中“手动配置”表中A列的标记

     二、运行：
        1.%s
        2.1.点“生成分类”按钮，运行成功后，程序同文件夹下，生成文件：%s
        2.2.点“合并文件”按钮，运行成功后，各源文件同文件夹下，生成文件：汇总-文件夹名.txt
     -------------------------------------------------------------------------------------
    ''' % (tip_txt_path,outputFilename)

    txt_info=Text(myFrame,height=6,relief='solid',padx=5,pady=5)
    txt_info.insert('end',str_info)

    # 第一行
    r=1
    myFrame.grid_rowconfigure(r,weight=1)

    btn_selectPath=Button(myFrame,text='选择文件夹')
    txt_path=Text(myFrame,height=2,relief='solid',bg='#FFFFF0',fg='#696969',padx=5,pady=5)
    btn_selectPath.config(command=lambda : [txt_path.delete('1.0','end'),txt_path.insert('0.0',filedialog.askdirectory().replace('/','\\'))])     # 设置点击事件

    # 内部函数
    def cancelTopmost(event):
        mainWindow.attributes('-topmost',False)     # 取消置顶
        # print('unbind:'+'<'+EventType(event.type).name+'>')
        event.widget.unbind('<'+EventType(event.type).name+'>')     # 解绑事件

    txt_path.bind('<FocusIn>',lambda event:cancelTopmost(event))    # 绑定焦点进入事件

    txt_path.focus_set()    # 获取焦点

    txt_path.insert('end',tip_txt_path)

    c=1
    btn_selectPath.grid(row=r,column=c,sticky='e')

    c+=1
    myFrame.grid_columnconfigure(c,weight=1)

    txt_path.grid(row=r,column=c,sticky='we')

    # 第二行
    r+=1
    c=1
    myFrame.grid_rowconfigure(r,weight=1)

    btn_classify=Button(myFrame,text='生成分类')
    btn_classify.config(command=lambda : classify(txt_path.get('0.0','end').strip(),outputFilename,txt_info))  # 设置点击事件
    btn_classify.grid(row=r,column=c,sticky='e')

    c+=1
    btn_merge=Button(myFrame,text='合并文件')
    btn_merge.config(command=lambda :merge(txt_path.get('0.0','end').strip(),txt_info))
    btn_merge.grid(row=r,column=c,sticky='e')

    # 第三行
    r+=1
    c=2
    myFrame.grid_rowconfigure(r,weight=5)

    txt_info.grid(row=r,column=1,columnspan=c,sticky='news')

    # 创建一个Scrollbar组件，并将它与Text组件绑定
    myScrollbar = Scrollbar(myFrame,command=txt_info.yview,orient='vertical')
    myScrollbar.grid(row=r,column=3,sticky='nes')

    # 将Text组件与Scrollbar组件进行关联
    txt_info.config(yscrollcommand=myScrollbar.set)

    myFrame.pack(fill='both',expand=True)

    mainWindow.mainloop()


def classify(dir_path,outputFilename,txt_info):
    try:
        txt_info.delete('1.0','end')
        if os.path.exists(dir_path):
            txt_info.insert('end','%s 开始分类...\n\n' % time.strftime('%Y-%m-%d %H:%M:%S'))
            txt_info.update()

            runRegExpFindAll(dir_path,outputFilename)

            txt_info.insert('end','%s √成功！\n\t生成的文件是程序同目录下的：%s' % (time.strftime('%Y-%m-%d %H:%M:%S'),outputFilename))

        else:
            txt_info.insert('end','%s ×路径不存在！' % time.strftime('%Y-%m-%d %H:%M:%S'))

    except PermissionError as e:
        txt_info.insert('end','%s ×错误：\n\t程序同目录下Excel已打开(config.xlxs或%s)，请关闭它后，重新运行！' % (time.strftime('%Y-%m-%d %H:%M:%S'),outputFilename))

    except Exception as e:
        str_error='×错误：\n%s' % traceback.format_exc()
        print(str_error)

        txt_info.insert('end','%s %s' % (time.strftime('%Y-%m-%d %H:%M:%S'),str_error))


def mergeContent(dir_path,txt_info):
    # root当前遍历到的文件夹路径；dirs当前文件夹中的所有文件夹名；files当前文件夹中的所有文件名
    for root, dirs, files in os.walk(dir_path):
        dict_startnumber={}

        # print(root)

        foldername=root[root.rfind('\\')+1:]
        filename_output='汇总-%s_python_zhangyu.txt' % foldername

        # 先清除一下输出的汇总文件
        outputPath='%s\\%s' %(root,filename_output)
        if os.path.exists(outputPath): os.remove(outputPath)    # 删除文件

        # print(root)
        txt_info.insert('end','%s\\\n' % foldername)
        txt_info.update()
        txt_info.yview('end')

        # # 在所有文件路径循环取出文本内容，获取文件名的开头
        for file in files:
            # print(file)
            path_sub=os.path.join(root, file)

            if path_sub==outputPath:continue

            # 匹配文件开头的数字
            match_num=re.match('\s*\d+\s*(?=\D)',file)

            if match_num is None:
                # 文件没有序号的都放在-1
                num_head=-1
            else:
                num_head=int(match_num.group())

            if num_head in dict_startnumber:
                dict_startnumber[num_head].append(path_sub)
            else:
                dict_startnumber[num_head]=[path_sub]

        # 如果字典为空，跳过
        if not dict_startnumber:continue

        # print(dict_startnumber)

        # 文件名排序
        li_startnumber=list(dict_startnumber.keys())
        li_startnumber.sort()   # 排序
        # print(li_startnumber)

        for n in li_startnumber:
            # 文件名没有序号的可能有多个，所以用list
            li_path_file=dict_startnumber[n]

            for path_file in li_path_file:
                # 获取源文本
                str_source=getStr_source(path_file)

                if str_source.strip()=='': continue

                filename_src=path_file.split('\\')[-1]

                # 追加到汇总文件
                with open(outputPath,'a',encoding='utf8') as f:
                    f.write('-- 来源于文件：%s\n%s\n\n' % (filename_src,str_source))

                str_info='\t读取\t%s\n\t写到\t%s\n\n'
                # print(str_info %(path_file,outputPath))
                txt_info.insert('end',str_info %(filename_src,filename_output))
                txt_info.update()
                txt_info.yview('end')


def merge(dir_path,txt_info):
    try:
        txt_info.delete('1.0','end')
        if os.path.exists(dir_path):
            txt_info.insert('end','%s 开始合并...\n\n' % time.strftime('%Y-%m-%d %H:%M:%S'))
            txt_info.update()

            mergeContent(dir_path,txt_info)

            txt_info.insert('end','%s √成功！\n\t生成的文件是各目录下的：汇总-文件夹名.txt' % time.strftime('%Y-%m-%d %H:%M:%S'))
        else:
            txt_info.insert('end','%s ×路径不存在！' % time.strftime('%Y-%m-%d %H:%M:%S'))
    except Exception as e:
        str_error='×错误：\n%s' % traceback.format_exc()
        print(str_error)
        txt_info.insert('end','%s %s' %(time.strftime('%Y-%m-%d %H:%M:%S'),str_error))


if __name__ =='__main__':
    windowInit()
