import pandas as pd
import time
import openpyxl
import traceback
import logging
import os


logging.basicConfig(level=logging.ERROR,    # 控制台打印的日志级别
            filename='log_error.txt',
            filemode='a',   # 模式，有w和a，w就是写模式，默认是追加模式
            format='%(asctime)s - [line:%(lineno)d] - %(levelname)s: %(message)s',  # 日志格式
            encoding='utf8')


def main():
    path_source='数据源'
    filename_dps='网点人员管理.xlsx'
    col_dlr='专营店'
    col_username='人员姓名'
    col_job_dps='人员岗位'
    col_dlr_user=col_dlr +'-' + col_username

    print('--处理DPS数据')

    print('%s 读取excel：DPS' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    df_dps=pd.read_excel(path_source+'/'+filename_dps,header=0)

    li_job=['岗位1','岗位2']

    print('%s 筛选' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    # 筛选
    df_dps=df_dps[df_dps['人员岗位'].isin(li_job)]

    print('%s 删除重复值' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    # 删除重复值
    df_dps.drop_duplicates(inplace=True)
    # print(df_dps)

    print('%s 拼接新列' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    # 拼接成新的列
    df_dps[col_dlr_user]=df_dps[col_dlr]+'-'+df_dps[col_username]

    # print(df_dps)

    print('--处理企业微信打卡数据')

    print('%s 读取excel：企业微信打卡' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    filename_wx='上下班打卡_日报.xlsx'

    df_wx=pd.read_excel(path_source+'/'+filename_wx,sheet_name='打卡详情',header=None)

    print('%s 删除空行空列' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    # 删除空行空列
    df_wx.dropna(axis=0,how='all',inplace=True)
    df_wx.dropna(axis=1,how='all',inplace=True)

    # 第3行作为列名的列表
    new_header=df_wx.iloc[2]

    print('%s 删除前3行' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    # 删除前3行
    df_wx.drop(df_wx.index[0:3],inplace=True)

    # 列名重命名
    df_wx.columns=new_header

    # # 重新设置索引
    # df_wx.reset_index(drop=True,inplace=True)
    # print(df_wx)

    col_date='日期'
    col_name='姓名'
    col_department='部门'
    col_job='职务'
    col_time_clock_in='实际打卡时间'

    li_cols=[col_date,col_name,col_department,col_job,col_time_clock_in]

    # 取部分列 
    df_wx_number_oneday=df_wx[li_cols].copy()

    series_department=df_wx_number_oneday[col_department]

    col_smallarea='小区'
    col_bigarea='大区'

    print('%s 部门分列成大区、小区、专营店' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    # 部门分列成大区、小区、专营店
    df_wx_number_oneday[col_dlr]=series_department.str.split('/').str[-1]
    df_wx_number_oneday[col_smallarea]=series_department.str.split('/').str[-2]
    df_wx_number_oneday[col_bigarea]=series_department.str.split('/').str[-3]

    print('%s 实际打卡时间列值替换--为0，其他为1' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))
    # 列值为--替换为0
    df_wx_number_oneday.loc[df_wx[col_time_clock_in]=='--',col_time_clock_in]=0
    df_wx_number_oneday.loc[df_wx[col_time_clock_in]!='--',col_time_clock_in]=1

    li_cols=[col_department,col_job,col_bigarea,col_smallarea,col_dlr,col_time_clock_in]
    dict_agg={}

    for li in li_cols:
        if li==col_time_clock_in:
            dict_agg[col_time_clock_in]='sum'
        else:
            dict_agg[li]='max'

    print('%s 分组合计：每天打卡数' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    # 分组合计
    df_wx_number_oneday=df_wx_number_oneday.groupby([col_name,col_date],as_index=False).agg(dict_agg)

    print('%s 是否打卡(0未打卡，1打卡)：每天是否打卡' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))
    df_wx_isclick_oneday=df_wx_number_oneday.copy()
    df_wx_isclick_oneday.loc[df_wx_number_oneday[col_time_clock_in]!=0,col_time_clock_in]=1

    # 列重命名
    df_wx_number_oneday.rename(columns={col_time_clock_in:'打卡次数'},inplace=True)

    df_wx_number_interval=df_wx_isclick_oneday.copy()

    # 获取日期列的唯一值
    num_days=len(df_wx[col_date].unique())

    print('%s 分组合计：%s天内打卡数' % (time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()) , num_days))
    df_wx_number_interval=df_wx_number_interval.groupby([col_name],as_index=False).agg(dict_agg)

    df_wx_isclick_oneday.rename(columns={col_time_clock_in:'是否打卡(0否，1是)'},inplace=True)

    df_wx_number_interval.rename(columns={col_time_clock_in:'%s天内打卡天数' % num_days},inplace=True)

    print('%s 明细分析保存到excel' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    # 创建一个ExcelWriter对象，with会自动关闭writer
    with pd.ExcelWriter('_detail_analysis.xlsx',engine='openpyxl') as writer:
        dict_df={
                # '每天打卡次数' : df_wx_number_oneday,
                '每天是否打卡' : df_wx_isclick_oneday,
                '%s天内打卡天数' % num_days : df_wx_number_interval}

        for sheet_name,df in dict_df.items():
            df.to_excel(writer,sheet_name=sheet_name,index=False)
            set_columns_width(writer,sheet_name,df)

    print('--合并连接DPS、企业微信打卡数据')

    print('%s DPS专营店-姓名、企微姓名追加到一列（专营店-姓名_ALL），去重' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))
    # DPS专营店-姓名、企微姓名追加到一列，去重
    series_user_all=pd.concat([df_dps[col_dlr_user],df_wx[col_name]],axis=0).drop_duplicates()

    col_dlr_user_all=col_dlr_user+'_ALL'
    df_dlr_user_all=pd.DataFrame(series_user_all,columns=[col_dlr_user_all])
    # print(df_dlr_user_all)

    # print('***********')

    print('%s 专营店-姓名_ALL、DPS左连接成临时表' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    suffixe='_DPS'
    df_dps.columns=[col+suffixe for col in df_dps.columns.tolist()]

    # 左连接
    df_result=pd.merge(df_dlr_user_all,
                        df_dps,
                        how='left',left_on=col_dlr_user_all,right_on=col_dlr_user+suffixe)

    # 删掉列
    df_result.drop(columns=[col_dlr_user+suffixe],inplace=True)

    print('%s 临时表、%s天打卡天数左连接成结果表' % (time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),num_days))
    suffixe='_企微'
    df_wx_number_interval.columns=[col+suffixe for col in df_wx_number_interval.columns.tolist()]

    df_result=pd.merge(df_result,
                        df_wx_number_interval,
                        how='left',left_on=col_dlr_user_all,right_on=col_name+suffixe)

    print('--保存结果')
    # print(df_result)
    print('%s 结果表保存到excel' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    sheet_name='result'
    filename_result=sheet_name+'.xlsx'
    with pd.ExcelWriter(filename_result,engine='openpyxl') as writer:
        df_result.to_excel(writer, index=False,sheet_name=sheet_name)
        set_columns_width(writer,sheet_name,df_result)

    print('%s 复制sheet数据到统计excel' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))
    path_filename='最后结果/打卡统计.xlsx'

    wb_src=openpyxl.load_workbook(filename_result,read_only=True)
    sheet_src=wb_src[sheet_name]

    wb_target=openpyxl.load_workbook(path_filename)
    sheet_target=wb_target[sheet_name]

    # 清空数据
    sheet_target.delete_rows(1,sheet_target.max_row)

    # 复制数据
    for row in sheet_src.iter_rows(values_only=True):
        sheet_target.append(row)

    # 刷新透视表
    for sht in wb_target._sheets:
        for pivot in sht._pivots:
            print('刷新 %s 中 透视表' % sht.title)
            pivot.cache.refreshOnLoad = True  # 根据最新数据刷新透视表

    wb_src.close()
    wb_target.save(path_filename)

    print('%s 打开统计excel' % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

    os.startfile(os.path.abspath(path_filename))

    print('--完成。')

    input('按任意键退出')


def set_columns_width(writer,sheet_name,df):
    # 获取openpyxl的工作表对象
    worksheet = writer.sheets[sheet_name]

    # 在列中循环设置列宽
    for i in range(df.shape[1]):
        c_letter=openpyxl.utils.get_column_letter(i+1)  # 列数字号变字母
        worksheet.column_dimensions[c_letter].width=20 # 设置列宽

    # # 冻结窗格
    # worksheet.freeze_panes='B2'


if __name__=='__main__':
    try:
        main()
    except Exception:
        str_error='×错误：%s' % traceback.format_exc()
        logging.error(str_error)
        input(str_error)
