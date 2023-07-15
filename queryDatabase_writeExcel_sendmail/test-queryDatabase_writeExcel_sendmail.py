import datetime

import shutil

# import os
import glob

import re

import ast

import pymysql

from email.mime.text import MIMEText
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
import smtplib

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Side,Border
from openpyxl.utils import get_column_letter


class DbMysql():
    def __init__(self, **arg):
        self.DBuser = arg.get('dbuser')
        self.PASSWORD = arg.get('pwd')
        self.Host = arg.get('host')
        self.Port = arg.get('port')
        self.DBname = arg.get('dbname')

    def conndb(self):
        try:
            self.db = pymysql.connect(host=self.Host, port=int(self.Port), user=self.DBuser, passwd=self.PASSWORD,
                                      db=self.DBname, charset='utf8mb4')

            # self.cursor = self.db.cursor(cursor=pymysql.cursors.DictCursor)    # 查询结果输出字典形式
            self.cursor = self.db.cursor()    # 默认tuple形式((),())

        except Exception as e:
            print('---connect error')
            print(e)

    def execute(self, sqlStr):
        try:
            self.cursor.execute(sqlStr)

            # 获取表头
            list_title = [str(i[0]) for i in self.cursor.description]

            # 获取数据
            tuple_data = self.cursor.fetchall()

            # print(tuple_data)

        except Exception as e:
            print('---execute error')
            print(e, flush=True)
            self.disconnect()

        return list_title,tuple_data

    def disconnect(self):
        try:
            self.cursor.close()
            self.db.close()
        except Exception as e:
            pass


# 函数：发送邮件
def send_mail(from_addr,mail_host,mail_port,mail_pass,subject,msg, list_To,list_Cc=[], list_attach_file_path=[]):
    message = MIMEMultipart()   # 多种格式

    # 邮件显示部分
    message['From'] = Header(from_addr)             # 发送邮箱
    message['To'] = Header(';'.join(list_To))       # 接收邮箱
    message['Subject'] = Header(subject, 'utf-8')   # 主题
    message['Cc'] = Header(';'.join(list_Cc))       # 抄送

    message.attach(MIMEText(msg, 'html', 'utf-8'))  # 内容

    # 添加附件
    if list_attach_file_path:
        for file_path in list_attach_file_path:
            filename=file_path.split('\\')[-1]
            print('    邮件附件：%s' % filename)

            f = MIMEApplication(open(file_path, 'rb').read())

            f['Content-Type'] = 'application/octet-stream'
            f.add_header('Content-Disposition','attachment', filename=filename)

            message.attach(f)

    if str(mail_port)=='465':
        smtpObj = smtplib.SMTP_SSL(host=mail_host, port=mail_port)  # smtp对象
    else:
        smtpObj = smtplib.SMTP(host=mail_host, port=mail_port)  # smtp对象

    # smtpObj.connect(mail_host, port=mail_port)
    #smtpObj = smtplib.SMTP(mail_host, port=mail_port)等同于smtpObj = smtplib.SMTP()、smtpObj.connect(mail_host, port=mail_port)两行代码

    # smtpObj.ehlo()
    # smtpObj.starttls()
    smtpObj.login(user=from_addr, password=mail_pass)

    smtpObj.sendmail(from_addr=from_addr, to_addrs=list_To+list_Cc, msg=message.as_string())    # 发送邮件

    smtpObj.quit()  # 退出

    print('发送邮件成功！')


def decryption(str_secret,salt):
    list_secret=str_secret.split('O')

    list_salt=salt.split('OO')

    len_prefix=len(str(int(list_salt[0],2)))

    str_indexes=list_salt[1]

    list_index=str_indexes.split('O')
    list_index_10=[int(i,2) for i in list_index]
    list_sort=list_index_10[:len_prefix]
    list_sort=sorted(list_sort)
    # print(list_sort)

    for i in range(len_prefix-1,-1,-1):
        index=int(list_sort[i])
        del(list_secret[index])

    list_char=[]

    for i in range(len(list_secret)):
        num=int(list_secret[i],16)+int(list_index_10[i])
        c=chr(num)
        list_char.append(c)

    str_password=''.join(list_char)

    return str_password


def create_workbook(list_title,tuple_data,filename):

    list_result=[]
    list_result.append(list_title)
    list_result.extend(tuple_data)

    # print(list_result)

    wb=Workbook()  # 创建工作簿
    sh=wb.worksheets[0]

    range_coloumns=range(len(list_title))
    list_column_width=[10 for i in range_coloumns]

    fille = PatternFill('solid', fgColor="C0C0C0")  # 填充色
    side = Side(style='thin', color='000000')   # 表格线样式
    border = Border(top=side, bottom=side, left=side, right=side)   #表格线

    r=0
    # 数据赋值给单元格
    for i in range(len(list_result)):
        r+=1

        if r == 1048577:
            # 大于excel最大行1048576时
            r=1
            sh=wb.create_sheet()    # 创建新sheet

        for j in range(len(list_result[i])):
            cell_data=sh.cell(r,j+1)

            cell_data.border=border    # 表格边框

            if r==1:
                # 表头
                value_data=list_result[0][j]
                cell_data.fill=fille   # 填充单元格颜色
            else:
                # 数据行
                value_data=list_result[i][j]

            cell_data.value=value_data

            # 获取列宽
            if not value_data is None:
                length_value =len(str(value_data).encode('gbk'))

                if length_value>list_column_width[j]:
                    list_column_width[j]=length_value

    # 设置列宽
    for i in range_coloumns:
        c_letter=get_column_letter(i+1)  # 列数字号变字母
        sh.column_dimensions[c_letter].width=list_column_width[i]+2

    wb.save(filename)
    sh=None
    wb=None


def run():
    
    sqlfile_path='sql_zhangyu'

    # 取出所有sql文件路径
    list_file_abs= glob.glob(sqlfile_path +'\\*.sql')

    list_sql=[]

    dict_param={}

    print('\n----------START----------\n')

    prefix_filename='Prefix_filename.sql'
    sql_prefix=''


    # 在所有sql文件路径中循环
    for file_abs in list_file_abs:
        with open(file_abs,'r',encoding='utf-8') as f:
            str_sql=f.read().strip()

        # 查出要输入的参数
        list_params=re.findall('@_(?!变量名)\w+_',str_sql)
        # print(list_params)

        filename=file_abs.replace(sqlfile_path+"\\",'')

        if list_params:
            # 参数列表不为空
            for param in list_params:
                if param not in dict_param:
                    param_value=input(' > 请输入 %s 中参数：%s=' % (filename,param)).strip()
                    dict_param[param]=param_value
                else:
                    param_value=dict_param[param]   # 取出字典中存在的参数值
                    print('同步设置 %s参数：%s=%s:' % (filename,param,param_value))

                str_sql=str_sql.replace(param,"'%s'" % param_value)

        list_sql.append(str_sql)

        if filename==prefix_filename: sql_prefix=str_sql

    print('\n---%s Running---\n' % datetime.datetime.now())

    with open('k_zhangyu','r') as f:
        list_key=f.readlines()

    dbconfMap['pwd']=decryption(list_key[0],list_key[1])

    db=DbMysql(dbuser=dbconfMap['dbuser'], pwd=dbconfMap['pwd'],
            host=dbconfMap['host'], port=dbconfMap['port'], dbname=dbconfMap['dbname'])
    db.conndb()

    prefix_excel=''

    # print(sql_prefix)

    if sql_prefix !='':
        print('查询Excel文件名前缀...')
        _,tuple_data=db.execute(sql_prefix)  # 查询数据库

        if tuple_data:
            prefix_excel=tuple_data[0][0]
        else:
            print('查不到Excel文件名前缀，请处理正确后，重新运行！\n')
            return

    excelfile_path='excel_zhangyu'

    with open(filename_msg_mail,'a',encoding='utf-8') as f:
        f.write(prefix_excel)

    # 在所有sql文件路径中循环
    for i in range(len(list_file_abs)):
        filename=list_file_abs[i].replace(sqlfile_path+'\\','')

        # Excel前缀文件时，跳过
        if filename==prefix_filename: continue

        str_sql=list_sql[i]

        if str_sql=='':
            print('%s 语句为空\n' % filename)
            continue

        print('\n%s 查询数据库...' % filename)

        # 查询数据库
        list_title,tuple_data=db.execute(str_sql)
        # print('%s\n%s' %(list_title,tuple_data))

        filename_noExtension=filename[:-4]
        if not tuple_data:
            print('%s 查询无数据\n' % filename)

            with open(filename_msg_mail,'a',encoding='utf-8') as f:
                f.write('，%s(数量：0)' % filename_noExtension)
            continue

        num=len(tuple_data)

        filename='%s\\%s-%s(数量：%s).xlsx' %(excelfile_path,prefix_excel,filename_noExtension,num)

        with open(filename_msg_mail,'a',encoding='utf-8') as f:
                f.write('，%s(数量：%s)' % (filename_noExtension,num))
        print('数据保存到：%s ...' % filename)

        # 保存数据到excel
        create_workbook(list_title,tuple_data,filename)

        list_title=[]
        tuple_data=()

    db.disconnect()

    db=None

    with open(filename_msg_mail,'a',encoding='utf-8') as f:
        f.write('\n')

    if input('\n > 是否发送邮件？（发送邮件请输入：y  不发送输入其他值）：').strip() in ['y','Y']:
        print('发邮件...')
        mail_pass=decryption(list_key[2],list_key[3])

        list_attach_file_path=glob.glob(excelfile_path +'\\*.xlsx')

        with open(filename_msg_mail,'r',encoding='utf-8') as f:
            msg=f.read()

        msg='%s：\n\n%s\n祥情见附件' % (subject,msg)
        msg=msg.replace('\n','<br/>').replace(' ','&nbsp')

        # 发送邮件
        send_mail(from_addr,mail_host,mail_port,mail_pass,subject,msg, list_To,list_attach_file_path=list_attach_file_path)

    print('\n------%s END------\n' % datetime.datetime.now())


if __name__ == '__main__':
    str_info='''
        /*-----------功能用途：为了查询数据库数据，保存到excel，并发送邮件包含excel附件-----------

            一、文件夹、文件说明：程序路径下，路径和名称不要更改，详情见 使用说明.txt

            二、运行前准备（若配置不变，可不用修改），详情见 使用说明.txt
                1.sql_zhangyu文件夹存放sql文件；2.sql文件中设置条件中的值为变量。

            三、运行中：
                1.自动执行sql_zhangyu文件夹下所有的sql，保存查询结果到excel
                2.发邮件时，自动把excel_zhangyu文件夹下所有excel文件作为附件发送
        ------------------------------------------------------------------------------------------*/
    '''

    print(str_info)

    # 清除发邮件正文信息
    filename_msg_mail='msg_mail.txt'
    with open(filename_msg_mail,'w',encoding='utf-8') as f:
        f.write('')

    excelpath='excel_zhangyu'
    list_path_excelfile=glob.glob('%s\\*.xlsx' % excelpath)

    historypath='excel_history'

    # excel移入历史表
    for file_abs in list_path_excelfile:
        filename=file_abs.replace(excelpath+'\\',historypath+'\\')
        shutil.move(file_abs,filename)

    # 读取配置
    with open('config.txt','r',encoding='utf-8') as f:
        list_config=f.readlines()

    dict_config={}
    for line in list_config:
        line=line.strip()
        index=line.find('=')
        if index !=-1:
            dict_config[line[:index].rstrip()]=line[index+1:].strip()

    # --设置公布变量--
    # 数据库参数
    dbconfMap=ast.literal_eval(dict_config['dbconfMap'])    # 字符串转字典

    # 邮件参数
    from_addr = dict_config['from_addr']
    mail_host = dict_config['mail_host']  # mail.dfpv.com.cn
    mail_port = dict_config['mail_port']  # 端口号：#587 #465 #25
    subject= dict_config['subject']
    list_To= ast.literal_eval(dict_config['list_To']) #ly-zhangyu@dfpv.com.cn,zhangcong@dfpv.com.cn,daishq@szlanyou.com
    # --设置公布变量

    while True:
        str_input=input(' > 开始运行请输入 y 退出请输入 exit：').strip()
        if str_input in ['y','Y']:
            run()
        elif str_input in ['exit','EXIT']:
            break
